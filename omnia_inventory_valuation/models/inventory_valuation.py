# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) 2004-2009 Tiny SPRL (<http://tiny.be>).
#    Copyright (C) 2010-2012 OpenERP s.a. (<http://openerp.com>).
#
#
#    Author : Smerghetto Daniel  (Omniasolutions)
#    mail:daniel.smerghetto@omniasolutions.eu
#    Copyright (c) 2014 Omniasolutions (http://www.omniasolutions.eu)
#    Copyright (c) 2018 Omniasolutions (http://www.omniasolutions.eu)
#    All Right Reserved
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################
'''
Created on Apr 17, 2018

@author: Matteo Boscolo
'''
from odoo import models
from odoo import fields
from odoo import api
from odoo import _
import logging
import datetime
from datetime import timedelta
from odoo.tools import DEFAULT_SERVER_DATETIME_FORMAT
from openpyxl import load_workbook
from odoo import _
import shutil
import stat
import tempfile
import os
import base64


class InventoryValuation(models.TransientModel):
    _name = "inventory.valuation"
    _description = 'Valuation of inventory at a defined date'

# descrizione in italiano
# Security

    ref_date = fields.Date(_('Date reference'))
    warehouse_id = fields.Many2one('stock.warehouse', _('Warehouse'))
    location_id = fields.Many2one('stock.location', _('Location'))
    show_zero = fields.Boolean(_('View zero quantity'))
    datas_fname = fields.Char(_('Optional Filename'))
    datas = fields.Binary(_('File content'))

    def filterLocationsByWarehouse(self, target_stock, locations):
        out = self.env['stock.location']
        
        def checkRecursion(target, location):
            if location.id == target.id:
                return True
            if not location.location_id:
                return False
            return checkRecursion(target, location.location_id)
            
            
        for location in locations:
            if location.id == target_stock.id:
                out += location
                continue
            if checkRecursion(target_stock, location):
                out += location
        return out

    def getLocations(self):
        location_ids = self.env['stock.location']
        if self.location_id:
            location_ids = self.location_id
        else:
            internal_locations = location_ids.search([('usage', '=', 'internal')])
            if not self.warehouse_id:
                location_ids = internal_locations
            else:
                location_ids = self.filterLocationsByWarehouse(self.warehouse_id.lot_stock_id, internal_locations)
        return location_ids

    def evaluateResults(self, location_ids, valuation_layer_ids):
        out_dict = {}
        for valuation_layer in valuation_layer_ids:
            location = valuation_layer.stock_move_id.location_dest_id
            if location in location_ids:
                product = valuation_layer.product_id
                categ_id = product.categ_id
                default_dict = {
                    'unit_costs': [],
                    'price': 0,
                    'qty': 0
                    }
                out_dict.setdefault(location, {})
                out_dict[location].setdefault(categ_id, {})
                out_dict[location][categ_id].setdefault(product, default_dict)
                
                if not self.show_zero:
                    if valuation_layer.unit_cost == 0:
                        continue
                out_dict[location][categ_id][product]['price'] += valuation_layer.value
                out_dict[location][categ_id][product]['qty'] += valuation_layer.quantity
                out_dict[location][categ_id][product]['unit_costs'].append(valuation_layer.unit_cost)
        return out_dict

    def action_generate_inventory(self):
        location_ids = self.getLocations()
        valuation_layer_obj = self.env['stock.valuation.layer']
        ref_date = self.ref_date or '5000-01-01'
        valuation_layer_ids = valuation_layer_obj.search([
            ('create_date', '<', ref_date)
            ])
        out_dict = self.evaluateResults(location_ids, valuation_layer_ids)
        if not out_dict:
            return
        template_file = os.path.join(os.path.dirname(__file__), 'inventory_valuation.xlsx')
        out_fname = self.datas_fname
        if not out_fname:
            out_fname = 'stock_inventory'
        if not out_fname.endswith('.xlsx'):
            out_fname += '.xlsx'
        self.datas_fname = out_fname
        full_path = os.path.join(tempfile.gettempdir(), out_fname)
        shutil.copy(template_file, full_path)
        os.chmod(full_path, stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
        workbook = load_workbook(filename=full_path)
        worksheetpivot = workbook['inventory']
        worksheetpivot.cell(row=1, column=2).value = self.warehouse_id.name if  self.warehouse_id else ''
        worksheetpivot.cell(row=2, column=2).value = self.ref_date or ''
        worksheetpivot.cell(row=3, column=2).value = self.location_id.name if self.location_id else ''

        worksheet = workbook['data_sheet']
        row_counter = 2
        prec = self.env['decimal.precision'].precision_get('Product Price')
        location_cache = {}
        category_cache = {}
        for loc, loc_vals in out_dict.items():
            for categ, categ_vals in loc_vals.items():
                for product_product, prod_vals in categ_vals.items():
                    if product_product.type != 'product':
                        continue
                    if prod_vals['price'] == 0 and not self.show_zero:
                        continue
                    location_name =  location_cache.get(loc)
                    if not location_name:
                        location_cache[loc] = loc.complete_name
                    categ_name  = category_cache.get(categ)
                    if not categ_name:
                        category_cache[categ] = categ.display_name
                    average_unit_cost = 0
                    if prod_vals['qty'] and prod_vals['unit_costs']:
                        average_unit_cost = sum(prod_vals['unit_costs']) / prod_vals['qty']
                    prod_name = product_product.display_name
                    worksheet.cell(row=row_counter, column=1).value = location_name
                    worksheet.cell(row=row_counter, column=2).value = categ_name
                    worksheet.cell(row=row_counter, column=3).value = prod_name
                    worksheet.cell(row=row_counter, column=4).value = round(average_unit_cost, prec)
                    worksheet.cell(row=row_counter, column=5).value = round(prod_vals['price'], prec)
                    worksheet.cell(row=row_counter, column=6).value = round(prod_vals['qty'], prec)
                    self.generate_inventory_line(worksheet, row_counter, product_product, location_name, categ)
                    row_counter += 1
        workbook.save(full_path)
        os.chmod(full_path, stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
        with open(full_path, 'rb') as f:
            fileContent = f.read()
            if fileContent:
                self.datas = base64.b64encode(fileContent)
        return {'name': _('Inventory Valuation'),
                'view_type': 'form',
                "view_mode": 'form',
                'res_model': 'inventory.valuation',
                'target': 'new',
                'res_id': self.ids[0],
                'type': 'ir.actions.act_window',
                'domain': "[]"}

    
    def generate_inventory_line(self, worksheet, row_counter, product_product, location_name, categ):
        return
